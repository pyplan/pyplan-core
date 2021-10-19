import importlib
import ntpath
import os
import re
import subprocess
import time

import numpy as np
import pandas as pd
import xarray as xr
from openpyxl import load_workbook

from .ws.settings import ws_settings

try:
    from StringIO import StringIO as BytesIO
except ImportError:
    from io import BytesIO


class PyplanFunctions(object):

    def __init__(self, model=None):
        self.model = model

    def release(self):
        self.model = None

    def set_domain(self, dataArray, domainDic, defaultValue=None):
        """Reindexes the dataArray by applying the indices of the domainDic param

           Ex. 
               pp.set_domain(da,{"time":time_idex, "products":product_index})
        """

        _da = dataArray
        for key in domainDic:
            _da = _da.reindex({key: domainDic[key].values})
            _da = _da.rename({key: domainDic[key].name})
        if not defaultValue is None:
            _da = _da.fillna(defaultValue)
        return _da

    def build_report(self, values, name="Report", report_index=None):
        """DEPRECATED. Use the create_report function instead
           Concatenates the values list of nodes along the report_index dimension
        """

        _titles = [str(xx.name) for xx in values]
        _index = None
        if report_index is None:
            _index = pd.Index(_titles, name=name)
        else:
            _index = report_index

        return xr.concat(values, _index)

    def create_dataarray(self, value, coords, dtype=None):
        """Creates a dataarray using an atomic value distributed along all dimensions 

           Ex. 
               pp.create_dataarray(1., coords=[time_idex, product_index])
        """

        _data = np.full(tuple([(len(x)) for x in coords]), value, dtype=dtype)
        return xr.DataArray(_data, coords)

    def find(self, param1, param2, compareType=1, caseSensitive=True):
        """
        param1: value or indexarray for compare
        param2: index compare to
        compareType: exact=1, start_with=2, end_with=3, contain=4  
        caseSensitive: able to differentiate between uppercase and lowercase (by default True)

        If param1 is a scalar (numeric or str) and param2 is an index:  return a dataArray indexed by param2 with True
        on ocurrences of param2
            Ex. pp.find("te", region, cp.end_with)
        If param1 is an index and param2 is an index too:  return a dataArray indexed by param1 and param2 with True
        on ocurrences of param1 on param2
            Ex. pp.find(subregion, region, cp.contain)
        """

        def _internalFn(item, value):
            if not isinstance(item, str):
                item = str(item)
            if not isinstance(value, str):
                value = str(value)

            if compareType == 1:
                if caseSensitive:
                    return item == value
                else:
                    return item.lower() == value.lower()
            elif compareType == 2:
                if caseSensitive:
                    return item[:len(value)] == value
                else:
                    return item[:len(value)].lower() == value.lower()
            elif compareType == 3:
                if caseSensitive:
                    return item[-len(value):] == value
                else:
                    return item[-len(value):].lower() == value.lower()
            elif compareType == 4:
                if caseSensitive:
                    return value in item
                else:
                    return value.lower() in item.lower()

        if (isinstance(param1, str) or str(param1).isnumeric()) and isinstance(param2, pd.Index):
            vfn = np.vectorize(_internalFn)
            return xr.DataArray(vfn(param2.values, param1), [param2])

        if isinstance(param1, pd.Index) and isinstance(param2, pd.Index):
            _res = self.create_dataarray(False, [param1, param2], dtype=bool)
            for row in param1.values:
                for col in param2.values:
                    _res.loc[{param1.name: slice(row, row), param2.name: slice(
                        col, col)}] = _internalFn(col, row)
            return _res

    def apply_fn(self, obj, applyFn, *args):
        """Applies "applyFn" to "obj" where obj can be DataArray or Index

           Ex.
               pp.apply(dataArray, node_function)
        """

        vfn = np.vectorize(applyFn)
        if isinstance(obj, pd.Index):
            return pd.Index(np.unique(vfn(obj.values, *args)))
        if isinstance(obj, xr.DataArray):
            return xr.apply_ufunc(vfn, obj, *args)
        return None

    def subset(self, cube):
        """Returns an index with the elements of the index for which cube is true. The function is used to create a new
           index that is a subset of an existing index.

           Ex. pp.subset(sales>0)
        """

        cond = cube > 0
        values = cond.coords[cond.dims[0]].values[cond.values]
        return pd.Index(values)

    def split_text(self, param1, separator, part=None):
        """Returns a DataArray object with text values formed by splitting the elements of param1 text values at each 
           occurrence of separator "separator". 
           The DataArray will have the original dimension plus a new dimension 'Parts' of length (number of separators + 1). 
           All text values must have the same number of separators separator.        
        """

        if isinstance(param1, pd.Index):
            param1 = xr.DataArray(param1.values, [param1])

        _q_separators = self.apply_fn(param1, lambda x: x.count(separator))
        _max_q_separators = np.asscalar(_q_separators.max().values)
        _result_coords = ['Part ' + str(i)
                          for i in range(1, _max_q_separators + 2)]
        _result_dim = pd.Index(_result_coords)
        _result_dim.name = "Parts"

        _results = []

        for _part in range(_max_q_separators + 1):
            _dataarray = self.apply_fn(
                param1, lambda x: x.split(separator)[_part])
            _results.append(_dataarray)

        _res = xr.concat(_results, dim=_result_dim)
        if not part is None:
            _res = _res.sel(Parts="Part " + str(part), drop=True)

        return _res

    def get_pos(self, index):
        """Returns a DataArray with indexed by index and its positions as values

           Ex. pp.get_pos(time_index)
        """

        return xr.DataArray(range(0, len(index)), [index])

    def concat_index(self, *args):
        """Concatenates two or more indexes and/or atomic values and returns a single new index

           Ex.
               pp.concatIndex(index1,index2,index3,value1,value2)
        """

        _list = []
        for arg in args:
            if isinstance(arg, pd.Index):
                values = (arg.values).tolist()
                _list.extend(values)
            else:
                _list.append(arg)

        seripandas = pd.Series(_list)
        return pd.Index(seripandas.unique())

    def linear_depreciation(self, investments, usefulLife, timeIndex, includeInCurrentMonth=False, timeIndexFormat='%Y.%m'):
        """Returns the straight-line depreciation of dataArray investments over its usefulLife.

           investments: DataArray containing investments
           usefulLife: DataArray with number of years of life expectancy
           timeIndex: Time dimension of dataArray. Must be a Pandas Index
           includeInCurrentMonth: Wheter to start depreciating in month t or month t+1
           timeIndexFormat: i.e. for '2016.01' would be '%Y.%m'
        """

        # Depreciation amount (safe division by zero)
        _usefulLife_months = usefulLife.astype(int) * 12
        _usefulLife_months_den = xr.where(
            _usefulLife_months == 0, 1, _usefulLife_months)
        _depreciation = xr.where(_usefulLife_months ==
                                 0, 0, investments / _usefulLife_months_den)

        # Calculate first and last months to depreciate
        _df_per = _usefulLife_months.to_dataframe('first').reset_index()
        _df_per['key'] = 1
        _df_time = pd.DataFrame(timeIndex)
        _df_time['key'] = 1
        _df = _df_time.merge(_df_per, on=['key']).drop(columns=['key'])
        _df['ending'] = pd.to_datetime(
            _df[timeIndex.name].str.replace('.', '-')).dt.to_period('M')
        _df['ending'] = (_df['ending'] + _df['first']
                         ).dt.strftime(timeIndexFormat)

        # Get dimensions indexes and names
        _getNodeFn = self.model.getNode

        _da_dims_names = list(usefulLife.dims)
        _da_dims = {timeIndex.name: timeIndex}
        for dim in _da_dims_names:
            _da_dims.update({dim: _getNodeFn(dim).result})

        # DataArray with ending date
        _ending = self.dataarray_from_pandas(
            _df, _da_dims, valueColumns='ending', defaultValue='')

        # Allocate depreciation to corresponding periods
        _depreciations = investments * 0.
        for t in timeIndex:
            _ending_month = self.subscript(_ending, timeIndex, t)
            _depreciation_amount_t = self.subscript(
                _depreciation, timeIndex, t)
            if includeInCurrentMonth:
                _depreciacion_t = xr.where((self.to_dataarray(timeIndex) >= t) & (
                    self.to_dataarray(timeIndex) < _ending_month), _depreciation_amount_t, 0.)
            else:
                _depreciacion_t = xr.where((self.to_dataarray(timeIndex) > t) & (
                    self.to_dataarray(timeIndex) <= _ending_month), _depreciation_amount_t, 0.)
            _depreciations = _depreciations + _depreciacion_t

        return _depreciations

    def irr(self, flow, time_index):
        """Returns the Internal Rate of Return (IRR) of a series of periodic payments (negative values) and inflows 
           (positive values). The IRR is the discount rate at which the Net Present Value (NPV) of the flows equals zero. 
           The variable flow must be indexed by time_index.

           If the cash flow never changes sign, pp.irr() has no solution and returns NAN (Not A Number).
        """

        _getNodeFn = self.model.getNode

        _rest_of_indexes_labels = np.setdiff1d(flow.dims, [time_index.name])

        _cube = None
        if len(_rest_of_indexes_labels) == 0:
            _cube = np.irr(flow.values)
        else:
            _rest_of_indexes = [_getNodeFn(
                xx).result for xx in _rest_of_indexes_labels]
            _cube = self.create_dataarray(0., _rest_of_indexes)
            _multivalues = [idx.values for idx in _rest_of_indexes]
            _values = pd.MultiIndex.from_product(_multivalues).values

            for _item in _values:
                _filter = {}
                for _nn in range(len(_item)):
                    _filter[_rest_of_indexes[_nn].name] = _item[_nn]
                _toIrr = flow.sel(_filter).values
                _irr = np.irr(_toIrr)
                _cube.loc[_filter] = _irr
        return _cube

    def copy_as_values(self, source, targetId):
        """Copy values of datArray "source" into dataArray with id 'targetId'. This function alters the definition of 
           dataArray with 'targetId' identifier.

           source: dataArray/index to copy values from
           targetId: identifier (string) of the target node 
        """

        _getNodeFn = self.model.getNode

        if isinstance(source, str):
            source = _getNodeFn(source).result

        if not isinstance(source, xr.DataArray) and not isinstance(source, pd.Index) and not isinstance(source, float) and not isinstance(source, int):
            raise ValueError(
                "The 'source' parameter must be a xr.DataArray, pd.Index, float or int")

        if not isinstance(targetId, str):
            raise ValueError(
                "The 'targetId' parameter must be a string (identifier of node)")

        newDef = ""
        if isinstance(source, float) or isinstance(source, int):
            newDef = f"result = {str(source)}"
        elif isinstance(source, xr.DataArray):
            _indexes = str(list(source.dims)).replace("'", '')
            np.set_printoptions(threshold=np.prod(source.values.shape))
            _data = np.array2string(source.values, separator=",", precision=20, formatter={
                                    'float_kind': lambda x: "np.nan" if np.isnan(x) else repr(x)}).replace('\n', '')
            newDef = f"result = xr.DataArray({_data},{_indexes})"
        elif isinstance(source, pd.Index):
            np.set_printoptions(threshold=np.prod(source.values.shape))
            _data = np.array2string(source.values, separator=",", precision=20, formatter={
                                    'float_kind': lambda x: "np.nan" if np.isnan(x) else repr(x)}).replace('\n', '')
            newDef = f"result = pd.Index({_data})"

        _getNodeFn(targetId).definition = newDef
        return True

    def excel_connection(self, filepath, useOpenpyxl=False, dataOnly=True, readOnly=True):
        """Creates Excel object from filepath.

           filepath: path to excel file
           useOpenpyxl: bool 
           dataOnly: bool. True to only get values, not the formula
           readOnly: bool
           Ex.
               pp.excel_connection("\\path\\to\\the\\excelfile.xlsx")
        """

        if self.model.isLinux():
            filepath = filepath.replace("\\", "/")

        _getNodeFn = self.model.getNode
        fullFilename = filepath

        if not os.path.isfile(fullFilename):
            fullFilename = _getNodeFn("current_path").result + filepath

        if os.path.isfile(fullFilename):
            if useOpenpyxl:
                return load_workbook(fullFilename, data_only=dataOnly, read_only=readOnly)
            else:
                return filepath
        else:
            raise ValueError("File not found")

    def subscript(self, dataArray, indexes, values):
        """Filters dataArray using the filterList filters. 

           dataArray: dataArray to be filtered
           indexes: the index to filter 
           values: the value to filter 
           Ex.
               pp.subscript(dataArray, index, value)
        """

        if not isinstance(dataArray, xr.DataArray):
            raise ValueError(
                "the 'dataArray' parameter must be of the type xr.DataArray")

        if not isinstance(indexes, list):
            indexes = [indexes]
        if not isinstance(values, list):
            values = [values]

        res = dataArray
        filterDic = {}
        for _pos, indexItem in enumerate(indexes):
            filterDic[indexItem.name] = values[_pos]

        if len(filterDic) > 0:
            res = res.sel(filterDic, drop=True)

        return res

    def change_index(self, dataArray, oldIndex, newIndex, compareMode=1, defaultValue=None):
        """Changes index of a dataArray object.

           compareMode: 1: by Value (default), 2: by pos
           Ex.
               pp.change_index(dataArray, oldIndex, newIndex)
        """

        _da = dataArray

        if compareMode == 1:
            _temp = _da.reindex({oldIndex.name: newIndex.values})
            _temp[newIndex.name] = _temp[oldIndex.name]
            _temp = _temp.swap_dims(
                {oldIndex.name: newIndex.name}).drop(oldIndex.name)
            if not defaultValue is None:
                _temp = _temp.fillna(defaultValue)
            return _temp
        else:

            if len(oldIndex.values) == len(newIndex.values):
                _tmp = _da.copy()
                _tmp = _tmp.assign_coords({oldIndex.name: newIndex.values})
                _tmp = _tmp.rename({oldIndex.name: newIndex.name})
                return _tmp
            elif len(oldIndex.values) > len(newIndex.values):
                raise ValueError(
                    "Changeindex by pos for indices of different size is not implemented")
            else:
                raise ValueError(
                    "Changeindex by pos for indices of different size is not implemented")

    def kind_to_string(self, kind):
        """Returns the data type on human-readable string
        """

        if kind in {'U', 'S'}:
            return "string"
        elif kind in {'b'}:
            return "boolean"
        elif kind in {'i', 'u', 'f', 'c'}:
            return "numeric"
        elif kind in {'m', 'M'}:
            return "date"
        elif kind in {'O'}:
            return "object"
        elif kind in {'V'}:
            return "void"

    def pandas_from_excel(self, excel, sheetName=None, namedRange=None, cellRange=None, indexes=None, driver=""):
        """Returns a pandas DataFrame from Excel spreadsheet.

           excel: excel file path or openpyxl workbook object
           sheetName: sheet name to be read
           namedRange: range name to be read
           cellRange: used together with sheetName to read from single cell range
           indexes: List of columns names to be set as index of dataframe
           Ex.
               pp.pandas_from_excel(excelNode,"Sheet 1")
               pp.pandas_from_excel(excelNode,namedRange="name_range")
               pp.pandas_from_excel(excelNode,"Sheet 1",cellRange="A1:H10")

           This function automatically generates pickles from every named range in excel file
           when excel parameter is a string.
        """

        # When excel param is a string, this function tries to read from automatically generated
        # pickles for every named range if they are newer than the Excel file (its modified date).
        # If they do not exist or are outdated, tries to generate one pickle for every named range in
        # the spreadsheet.
        # Requirements:
        #   - it must have writing permissions,
        #   - it must have named ranges.
        # Otherwise, it should load the spreadsheet using openpyxl library and then read the sheet,
        # range or cellrange.

        if isinstance(excel, str):
            if not os.path.isfile(excel):
                excel = os.path.join(self.model.getNode(
                    "current_path").result, excel)
            filepath = excel

            # Only read/generate pickles for named ranges
            if namedRange is not None:
                orig_dir, single_filename = os.path.split(filepath)
                filename, _ = os.path.splitext(single_filename)
                target_dir = os.path.join(orig_dir, f".{filename}")
                picklepath = os.path.join(target_dir, f"{namedRange}.pkl")

                # Read from pickle if it is newer than Excel file
                if os.path.isfile(picklepath) and os.path.getmtime(picklepath) >= os.path.getmtime(filepath):
                    return self.__read_pickle_df(filepath=picklepath, indexes=indexes)

                else:
                    wb = load_workbook(
                        filepath, data_only=True, read_only=True)
                    named_ranges = [
                        r.name for r in wb.defined_names.definedName]

                    # Check if user has writing permissions to generate new pickles and if namedRange exists
                    if os.access(excel, os.W_OK) and namedRange in named_ranges:
                        flag_filename = 'flag.tmp'
                        flag_filepath = os.path.join(target_dir, flag_filename)

                        # Clean potentially old flag files
                        self.__remove_old_file(
                            filepath=flag_filepath, maxElapsedMinutes=60)

                        # If flag file exists (optimization is running), read directly from Excel
                        if os.path.isfile(flag_filepath):
                            return self.pandas_from_excel(wb, sheetName, namedRange, cellRange, indexes)
                        else:
                            self.__generate_pkl_from_excel(
                                workbook=wb, filepath=filepath, targetDir=target_dir,
                                maxFileSizeMB=100, flagFilename=flag_filename)
                            # Read file
                            if os.path.isfile(picklepath):
                                return self.__read_pickle_df(filepath=picklepath, indexes=indexes)
                            else:
                                return self.pandas_from_excel(wb, sheetName, namedRange, cellRange, indexes)

                    # Read directly from Excel
                    else:
                        return self.pandas_from_excel(wb, sheetName, namedRange, cellRange, indexes)
            else:
                wb = load_workbook(filepath, data_only=True, read_only=True)
                return self.pandas_from_excel(wb, sheetName, namedRange, cellRange, indexes)

        elif "openpyxl.workbook" in str(type(excel)):
            rangeToRead = None
            if not namedRange is None:
                the_range = excel.defined_names[namedRange]
                dests = the_range.destinations
                for title, coord in dests:
                    ws = excel[title]
                    rangeToRead = ws[coord]
            elif not cellRange is None:
                ws = excel[sheetName]
                rangeToRead = ws[cellRange]
            else:
                rangeToRead = excel[sheetName]

            cols = []
            values = []
            for index, row in enumerate(rangeToRead):
                if index == 0:
                    cols = [str(c.value) for c in row]
                else:
                    values.append([c.value for c in row])
            df = pd.DataFrame(values, None, cols)
            if not indexes is None:
                if isinstance(indexes, str):
                    indexes = [indexes]
                toIndex = []
                for indexColumn in indexes:
                    if indexColumn in df.columns.values:
                        toIndex.append(indexColumn)
                if len(toIndex) > 0:
                    df.set_index(toIndex, inplace=True)
            return df.dropna(how="all")
        else:
            raise ValueError("excel must be a string or openpyxl workbook")

    def index_from_pandas(self, dataframe, columnName=None, removeEmpty=True):
        """Returns a pandas.Index from an column of a pandas dataframe.

           dataframe: pandas dataframe
           columnName: dataframe column name used for create cp.index. By default is created using the first column
           removeEmpty: True for remove empty rows
           Ex.
               pp.index_from_pandas(df)
               pp.index_from_pandas(df,"column10")
        """

        _serie = None
        if columnName is None:
            _serie = dataframe[dataframe.columns[0]]
        else:
            _serie = dataframe[columnName]

        if removeEmpty:
            _serie.dropna(inplace=True)
            if self.kind_to_string(_serie.dtype.kind) == "string" or self.kind_to_string(_serie.dtype.kind) == "object":
                _serie = _serie[_serie != ""]

        return pd.Index(_serie.unique())

    def index_from_excel(self, excel, sheetName=None, namedRange=None, cellRange=None, columnName=None, removeEmpty=True):
        """Returns a pandas.Index from an excel file.

           excel: pp.excel object
           sheetName: sheet name to be read
           namedRange: name of the range to be read
           cellRange: used with sheetname, for read from a simple range
           columnName: dataframe column name used for create pp.index. By default is created using the first column
           removeEmpty: True for remove empty rows
           Ex.
               pp.index_from_excel(excelNode,"Sheet 1")
               pp.index_from_excel(excelNode,namedRange="name_range")
               pp.index_from_excel(excelNode,namedRange="name_range", columnName="indicadores")
        """

        if isinstance(excel, str) or "openpyxl.workbook" in str(type(excel)):
            _df = self.pandas_from_excel(
                excel, sheetName, namedRange, cellRange)
            return self.index_from_pandas(_df, columnName, removeEmpty)
        else:
            raise ValueError(
                "excel can be excel_connection object or a str path to the filename")

    def dataarray_from_pandas(self, dataframe, domainDic, valueColumns, defaultValue=None, valueColumnsAsDim=True, sumDuplicateRecords=True):
        """Returns a DataArray (valueColumns is string or (valueColumns is pd.Index and valueColumnsAsDim is True)) 
           or Dataset (valueColumns is a list or (valueColumns is a pd.Index and valueColumnsAsDim is False)) from
           a Pandas dataframe applying the set_domain function.

           dataframe: Pandas dataframe with no index columns.
           domainDic: Dictionary of column names and index names. Ex. {'Column Name': index_name}.
           valueColumns: String, list or pd.Index. Dataframe's value columns.
           defaultValue: Default value when applying set_domain function.
           valueColumnsAsDim: If True, valueColumns becomes a dimension of resulting DataArray. If False, each value
           column becomes a variable of the resulting Dataset.
           sumDuplicateRecords: If True, sums identical rows. Otherwise, removes duplicates (except the first one). 
           Ex. 
               pp.dataarray_from_pandas(sales_dataframe, {'Sales Channel': sales_channels, 'Month': time}, 'Sales', 0.)
        """

        _index_value_columns = None

        # Check if valueColumns is string, list, np.ndarray or pd.Index (transform to list) and indexes is dict.
        if isinstance(valueColumns, pd.Index):
            _index_value_columns = valueColumns.copy()
            _index_value_columns_name = _index_value_columns.name
            valueColumns = valueColumns.values.tolist()
        elif isinstance(valueColumns, np.ndarray):
            valueColumns = valueColumns.tolist()
        elif not isinstance(valueColumns, str) and not isinstance(valueColumns, list):
            raise ValueError(
                "valueColumns must be a string, a list or a pd.Index")
        if not isinstance(domainDic, dict):
            raise ValueError("Indexes must be a dictionary")

        # Transform indexes into list and create list with all columns.
        _index_cols = list(domainDic.keys())
        _cols = _index_cols.copy()

        if isinstance(valueColumns, list):
            _cols = _cols + valueColumns
        else:
            _cols.append(valueColumns)

        # If valueColumnsAsDim is True, check if every column is in dataframe and filter it.
        if (valueColumnsAsDim is True) and isinstance(_index_value_columns, pd.Index):
            _df_columns = dataframe.columns.values.tolist()
            _cols = [value for value in _df_columns if value in _cols]
            _filtered_value_columns = [
                value for value in _cols if value not in _index_cols]

        # Filter dataframe by columns.
        _df = dataframe[_cols]

        # Sum identical rows or remove duplicates.
        if sumDuplicateRecords is True:
            _df = _df.groupby(_index_cols, as_index=False).sum()
        else:
            _duplicate_rows = _df.duplicated(_index_cols)
            _df = _df[~_duplicate_rows]

        # If valueColumnsAsDim is True, melt valueColumns.
        if (valueColumnsAsDim is True) and isinstance(_index_value_columns, pd.Index):
            # Unpivot dataframe from wide format to long format by valueColumns.
            _df = pd.melt(_df, id_vars=_index_cols, value_vars=_filtered_value_columns,
                          var_name=_index_value_columns_name, value_name='values')
            _index_cols = _index_cols + [_index_value_columns_name]
            domainDic[_index_value_columns_name] = _index_value_columns

            # Create DataArray
            _data = _df.set_index(_index_cols)['values'].to_xarray()

            # Appy set_domain function to DataArray / Dataset.
            _data = self.set_domain(_data, domainDic, defaultValue)
        else:
            # Create DataArray / Dataset.
            _data = _df.set_index(_index_cols)[valueColumns].to_xarray()

            # Appy set_domain function to DataArray / Dataset.
            _data = self.set_domain(_data, domainDic, defaultValue)

        return _data

    def dataarray_from_excel(self, excel, sheetName=None, namedRange=None, cellRange=None, indexes=None, valueColumns=None, indexColumnHeaders=None, replaceByIndex=None, defaultValue=0):
        """Returns a xr.DataArray from excel file.

           excel: excel_connection object.
           sheetName: sheet name to be read
           namedRange: name of the range to be read.
           cellRange: used with sheetName to read from a simple range.
           indexes: pd.Index objects to perform a change_index operation.
           valueColumns: string with the column name of the dataframe that contains the values. pd.Index with column
           names to convert columns to index.
           indexColumnHeaders (optional): column names in pandas to parse with indexes. Used if header on dataframe 
           is not equal to index identifiers.
           replaceByIndex (optional): replace index used in valueColumns by this index (using change_index).
           Ex.
               pp.dataarray_from_excel(excelNode,"Sheet 1",indexes=[indicadores],valueColumns="descuentos")
               pp.dataarray_from_excel(excelNode,namedRange="nombre_rango",indexes=[indicadores],valueColumns=time)
        """

        dataframe = self.pandas_from_excel(
            excel, sheetName, namedRange, cellRange)
        # Check size of dataframe. If it is empty, create empty dataArray. Else, proceed
        if len(dataframe) == 0:
            if not isinstance(indexes, list):
                indexes = [indexes]
            if isinstance(valueColumns, pd.Index):
                indexes.append(valueColumns)

            _data = np.full(tuple([(len(x)) for x in indexes]), defaultValue)
            return xr.DataArray(_data, indexes)
        else:
            valueIndex = None
            if isinstance(valueColumns, pd.Index):
                valueIndex = valueColumns
                valueColumns = valueIndex.tolist()
            elif isinstance(valueColumns, str):
                valueColumns = [valueColumns]

            if indexColumnHeaders is None:
                indexColumnHeaders = [index.name for index in indexes]

            # Create total index and index names
            _allindexes = indexes
            _allIndexNames = indexColumnHeaders[:]
            if not valueIndex is None:
                _allindexes.append(valueIndex)
                _allIndexNames.append("data_index")

            # fill other columns for prevent melt error
            cols_not_in_df = [
                col for col in valueColumns if col not in dataframe.columns]
            for col in cols_not_in_df:
                dataframe[col] = np.nan

            _full = dataframe.reset_index()[indexColumnHeaders + valueColumns].melt(
                id_vars=indexColumnHeaders,
                value_vars=valueColumns,
                var_name="data_index",
                value_name="data_value")

            # sum for acum over duplicate records
            _full = _full.groupby(_allIndexNames, as_index=False).sum()
            _dtype = _full["data_value"].dtype

            _dataType = self.kind_to_string(_dtype.kind)
            if _dataType == "string":
                _full = _full[(_full["data_value"] != "") &
                              (_full['data_value'].notna())]
            else:
                _full = _full[(_full["data_value"] != 0) &
                              (_full['data_value'].notna())]

            _full.set_index(_allIndexNames, inplace=True)
            _da = _full["data_value"].to_xarray()

            # If indexed, rename index
            if not indexes is None and not indexColumnHeaders is None:
                if not isinstance(indexes, list):
                    indexes = [indexes]
                idxPos = 0
                for cubeIndex in indexes:
                    newIndexName = cubeIndex.name
                    if idxPos <= len(indexColumnHeaders)-1:
                        oldIndexName = indexColumnHeaders[idxPos]
                        if not newIndexName in _da.coords:
                            _da.coords[newIndexName] = _da.coords[oldIndexName]
                            _da = _da.swap_dims(
                                {oldIndexName: newIndexName}).drop(oldIndexName)
                        idxPos += 1
                        # Reindex to complete combinations
                        _da = _da.reindex({newIndexName: cubeIndex.values})

            if not valueIndex is None:
                newIndexName = valueIndex.name
                oldIndexName = "data_index"
                if not newIndexName in _da.coords:
                    _da.coords[newIndexName] = _da.coords[oldIndexName]
                    _da = _da.swap_dims(
                        {oldIndexName: newIndexName}).drop(oldIndexName)
                # Reindex to complete combinations
                _da = _da.reindex({newIndexName: valueIndex.values})

                if not replaceByIndex is None:
                    _da = self.change_index(_da, valueIndex, replaceByIndex, 2)

            return _da.fillna(defaultValue)

    def to_dataarray(self, index):
        """Converts an index into DataArray indexed by index and with its values
           Ex.
               pp.to_dataarray(time_index)
        """

        return xr.DataArray(index.values, [index])

    def goal_seek(self, nodeIdX, nodeIdObjective, goal=0, startValue=1, matrixIndex=None):
        """Finds the value of nodeIdX that makes nodeIdObjective equal to goal.

           nodeIdX: String with id of node X
           nodeIdObjective: String with id of node Objective
           matrixIndex: Index for multidimensional goal seek
        """

        _getNodeFn = self.model.getNode
        if self._exists_module("scipy"):
            from scipy.optimize import newton

            if matrixIndex is None:
                def _f(x):
                    _getNodeFn(nodeIdX).definition = "result = " + str(x)
                    value = _getNodeFn(nodeIdObjective).result
                    return value - goal
                _res = newton(_f, x0=startValue)
                return _res
            else:
                _indexName = matrixIndex.name
                for item in matrixIndex:
                    def _f(x):
                        _values = _getNodeFn(nodeIdX).result
                        _values.loc[{_indexName: slice(item, item)}] = x
                        np.set_printoptions(
                            threshold=np.prod(_values.values.shape))
                        data = np.array2string(_values.values, separator=",", precision=20,
                                               formatter={
                                                   'float_kind': lambda x: "np.nan" if np.isnan(x) else repr(x)}
                                               ).replace('\n', '')
                        _getNodeFn(
                            nodeIdX).definition = f"result = xr.DataArray({data},[{_indexName}])"
                        value = _getNodeFn(nodeIdObjective).result
                        return self.subscript(value, matrixIndex, item)
                    _res = newton(_f, x0=startValue)
        else:
            raise ValueError("scipy library not found")

    def _exists_module(self, import_name):
        """Return true if module is installed
        """

        try:
            importlib.import_module(import_name)
            return True
        except ImportError:
            return False

    def install_library(self, pypi_name, import_name=None):
        """DEPRECATED. Use Lib manager instead
        """

        if import_name is None:
            import_name = pypi_name

        if not self._exists_module(import_name):
            # check in lib folder
            # install lib
            os.system(f"pip install {pypi_name}")
            importlib.invalidate_caches()
            if not self._exists_module(import_name):
                raise ValueError(f"Can't install the module '{import_name}'")
        return True

    def create_time(self, date_start, date_end, freq='M', format='%Y.%m'):
        """Creates time index usign start, end dates and freq. The result is formated with format parameter
           Ex.
               pp.create_time('2016.01','2018.12')
               pp.create_time('2016.01.01','2016.12.31',freq='D',format='%d/%m/%Y')
        """

        if "." in date_start:
            date_start = date_start.replace('.', '-')
        if "." in date_end:
            date_end = date_end.replace('.', '-')
        return pd.Index(pd.period_range(start=date_start, end=date_end, freq=freq).strftime(format))

    def lookup(self, dataArray, dataMap, sharedIndex, defaultValue=0):
        """Returns the value of dataArray indexed by the index of dataMap.

           dataArray must be indexed by sharedIndex and dataArray values must correspond to elements of sharedIndex.

           For example: Let's say you have a dataArray with an estimated inflation rate by Country ("inflation_rate"
           is the name of the dataArray; "country" is the name of the index) and you want to assign it to the 
           corresponding Company depending on its location. On the other hand, there's a many-to-one map where each
           Company is allocated to a single Country ("country_to_company_allocation"). The sharedIndex, in this case,
           is Country ("country").
           As a result, 
                pp.lookup( inflation_rate , country_to_company_allocation , country )
           will return the estimated inflation rate by Company.
        """

        try:
            return dataArray.sel({sharedIndex.name: dataMap}, drop=True)
        except:
            valuesOk = dataMap[dataMap.isin(sharedIndex.values)]
            lookOk = dataArray.sel({sharedIndex.name: valuesOk}, drop=True)
            final = lookOk.reindex(
                {dataMap.dims[0]: dataMap.coords[dataMap.dims[0]].values})
            return final.fillna(defaultValue)

    def aggregate(self, dataArray, mapInfo, sourceIndex, targetIndex, aggregationFunction='sum'):
        """Converts dataArray, originally indexed by sourceIndex, to a dataArray indexed by targetIndex, aggregating
           according to the mapInfo‘s allocation of targetIndex: sourceIndex.

           mapInfo: gives the value of targetIndex for each element of sourceIndex (If the map does not match then the
           element will not be set into target index and information will be lost)
           aggregationFuction (optional): especifies the function to be used when grouping data (sum, mean, min, max,
           median)

           Ex. for aggregating time information into annual index, the syntax is:
               pp.aggregate(dataArray, timeToYearsMap, time, years)
        """

        # Transform map and targetIndex to list
        if not isinstance(mapInfo, list):
            mapInfo = [mapInfo]
        if not isinstance(targetIndex, list):
            targetIndex = [targetIndex]

        if len(mapInfo) == len(targetIndex):
            # Create dataframe map with new indexes
            _map = pd.DataFrame(columns=[sourceIndex.name]).set_index(
                sourceIndex.name)
            for i in range(len(mapInfo)):
                _map_i = mapInfo[i].to_dataframe(targetIndex[i].name)
                _map = _map.join(_map_i, how='outer')

            _df = dataArray.to_dataframe('value')
            _empty_filter = _df["value"] != 0
            # Drop rows with 0 if dataframe is not empty (to avoid error)
            if len(_df[_empty_filter]) != 0:
                _df = _df[_empty_filter]

            # Join new dimensions (target) and drop original (source)
            _df = _df.join(_map).reset_index()
            _df.drop(columns=[sourceIndex.name], inplace=True)
            _newDimList = [
                xx for xx in dataArray.dims if xx not in [sourceIndex.name]]

            # Groupby dataframe by new dimensions
            for i in range(len(targetIndex)):
                _newDimList.append(targetIndex[i].name)
            _df = _df.groupby(_newDimList).agg(aggregationFunction)

            # Transform to Xarray DataArray
            _da = _df["value"].to_xarray()

            # Reindex dimensions
            _reindexDic = {}
            for t_index in targetIndex:
                _reindexDic.update({t_index.name: t_index.values})
            for coord in dataArray.coords:
                if coord != sourceIndex.name:
                    _reindexDic[coord] = dataArray.coords[coord].values
            _da = _da.reindex(_reindexDic)

            return _da.fillna(0)
        else:
            raise ValueError(
                'mapInfo and targetIndex must have the same number of elements')

    def choice(self, index, selection, includeAll=False):
        """DEPRECATED: Use pp.selector instead. 
           Returns the element in the "selection" position of the index. 
        """
        if selection == 0 and includeAll == 1:
            return "All"
        else:
            values = None
            if isinstance(index, pd.Index):
                values = (index.values[:1000]).tolist()
            elif isinstance(index, np.ndarray):
                values = (index[:1000]).tolist()
            else:
                values = list(index)[:1000]
            if not values is None and len(values) >= selection:
                return values[selection-1]
        return ""

    def dynamic(self, dataArray, index, shift, initialValues=None, sliceInputs=False):
        """Performs cyclic calculations between nodes.

           dataArray: dataArray to perform the ciclyc dependency calculation
           index: Index from dataArray to shift 
           shift: number of elemnts to shift. Can be positive or negative
           initialValues (optional): scalar or 1-dim dataArray. Initial values to apply to first "shift" elements
           sliceInputs (optional): boolean to apply subscript to every input in dynamic loop. If one of the nodes
            in dynamic loop is set to True, it applies it to every other node in the loop
        """

        _da = dataArray.shift({index.name: (shift*-1)})
        if not initialValues is None:
            _da = _da.fillna(initialValues)
        return _da

    def slice_dataarray(self, dataArray, index, position):
        """Filters dataArray by integer position along the specified index.

           dataArray: dataArray to be filtered
           index: pp.index 
           position: int 
           Ex.
               pp.slice_dataarray(dataArray1, index1, 0)
        """

        if not isinstance(dataArray, xr.DataArray):
            raise ValueError(
                "the 'dataArray' parameter must be of the type xr.DataArray")
        return dataArray.isel({index.name: position}, drop=True)

    def fill_inf(self, dataArray, value=0):
        """Fills np.inf values with value

           Ex.
               pp.fill_inf(dataArray, 0)
        """

        return self.apply_fn(dataArray, lambda x: value if np.isinf(x) else x)

    def fill_all(self, dataArray, value=0):
        """Fills np.inf and np.nan with value
           Ex.
               pp.fill_all(dataArray, 0)
        """

        return self.fill_inf(dataArray.fillna(value), value)

    def add_periods(self, start, periods, freq='M', format='%Y.%m'):
        """Adds periods to a date. Allows setting freq and output format 

           Ex.
                pp.addPeriods('2016.01', 6)
                pp.apply_fn(pp.addPeriods, projects_initial_date, projects_duration)
        """

        if "." in start:
            start = start.replace('.', '-')
        if periods < 0:
            return pd.period_range(end=start, periods=-periods+1, freq=freq).strftime(format)[0]
        else:
            return pd.period_range(start=start, periods=periods+1, freq=freq).strftime(format)[-1]

    def npv(self, rate, flow, time_index, offset=1):
        """"Returns the Net Present Value (NPV) of a cash flow with equally spaced periods. The flow parameter must contain
            a series of periodic payments (negative values) and inflows (positive values), indexed by time_index.
            The optional offset parameter especifies the offset of the first value relative to the current time period. By 
            default, offset is set to 1, indicating that the first value is discounted as if it is one step in the future
        """

        _number_of_periods = self.get_pos(time_index) + offset
        _present_values = flow / (1 + rate) ** _number_of_periods
        _npv = _present_values.sum(time_index.name)
        return _npv

    def copy_index(self, dataArray, sortValues=True):
        """Generates a pd.Index with the unique values of the dataArray.
        """

        np_values = dataArray.values.flatten()

        # Numpy unique function automatically reorders. Pandas unique, does not.
        if sortValues is True:
            return pd.Index(np.unique(np_values))
        else:
            return pd.Index(np_values).unique()

    def sequence_index(self, _start, _end, _step=1):
        """
        Returns a pd.Index with the sequence between 'start' and 'end' parameters. Both limits are inclusive. Values are 
        converted to string.
        """
        try:
            _start = int(_start)
            _end = int(_end) + 1
            _step = int(_step)
        except:
            raise ValueError(
                "Only numbers are allowed as 'start', 'end' and 'step' parameters")

        _list = [str(x) for x in range(_start, _end, _step)]
        _index = pd.Index(_list)
        return _index

    def subindex(self, dataArray, targetValue, targetIndex, method='Last'):
        """Returns a dataArray containing the value of targetIndex for which dataArray (indexed by targetIndex) is equal 
           to targetValue.

           dataArray: Xarray dataArray.
           targetValue: Integer, Float or String.
           targetIndex: Pandas Index.
           method: There are two options: "Last" returns the last occurrence of targetIndex for which dataArray is equal
           to targetValue. "First" returns the first occurrence.
        """

        # Equals dataArray to targetValue and cumulates it along targetIndex.
        _matriz_1_0 = xr.where(dataArray == targetValue, 1, 0)
        _matriz_1_0_acum = xr.where(
            _matriz_1_0 == 1, _matriz_1_0.cumsum(targetIndex.name), 0)

        if method == 'Last':
            # Get max cumulated value along targetIndex
            _max = _matriz_1_0_acum.max(targetIndex.name)
            _max = xr.where(_max == 0, np.nan, _max)
            _matriz_max = xr.where(
                _matriz_1_0_acum == _max, self.to_dataarray(targetIndex), np.nan)
            return _matriz_max.max(targetIndex.name)
        elif method == 'First':
            # Get min (1) cumulated value along targetIndex
            _matriz_min = xr.where(_matriz_1_0_acum == 1,
                                   self.to_dataarray(targetIndex), np.nan)
            return _matriz_min.max(targetIndex.name)
        else:
            raise ValueError("Insert a valid method")

    def concat_rows(self, array_param, index_param):
        """Flattens array_param by replacing with a new index that includes all combinatios of values from
           index_param
        """

        _index = pd.Index([])
        for i in index_param.values:
            _index = self.concat_index(_index, pd.Index(
                array_param.sel({index_param.name: i}, drop=True).values))
        return _index

    def log_task(self, task_state="PROGRESS", task_description=None, task_activity=None, task_info=None):
        """Generates log entry. Used for schedule tasks

           task_state: PROGRESS, INFO, WARNING, FAILURE, RETRY, SUCCESS, REVOKED, STARTED, PENDING, RECEIVED
           task_description: Shot description of task. example: start process
           task_activity: other short description
           task_info: json with more info 
        """

        import json
        _params = {
            "state": task_state,
            "description": task_description,
            "activity": task_activity,
            "info": json.dumps(task_info)}

        res = None

        task_log_endpoint = self.model.getNode("task_log_endpoint").result
        if task_log_endpoint:
            # only used from pyplan_engine
            import requests
            from os import environ
            base_host = environ['PYPLAN_API_HOST'] + task_log_endpoint
            res = requests.post(base_host, data=_params)
        else:
            print(str(_params))

        return res

    def pandas_from_xlsb_file(self, filepath):
        """Returns a pandas DataFrame from xlsb file
        """

        if self._exists_module("pyxlsb"):
            from pyxlsb import open_workbook as open_xlsb
            _df = []
            with open_xlsb(filepath) as wb:
                with wb.get_sheet(1) as sheet:
                    for row in sheet.rows():
                        _df.append([item.v for item in row])

            return pd.DataFrame(_df[1:], columns=_df[0])
        else:
            raise ValueError("pyxlsb library not found")

    def selector(self, options, selected, multiselect=False):
        """Creates UI Pyplan selector for decision nodes

           options: List or pandas.Index with values that can be selected 
           selected: current selected index value
           multiselect: True to allow multiple selection
        """

        return Selector(options, selected, multiselect)

    def send_message(self, message_text, message_title=None, not_level_reverse="info"):
        """Sends message to UI. Only used with Pyplan UI
           Ex.
               pp.send_message("The process has been completed","Process complete!","success")            
        """

        if self.model and self.model.ws:
            not_level = ws_settings.NOTIFICATION_LEVEL_CHOICES_REVERSE[
                not_level_reverse] if not_level_reverse in ws_settings.NOTIFICATION_LEVEL_CHOICES_REVERSE else ws_settings.NOTIFICATION_LEVEL_INFO
            self.model.ws.sendMsg(message_text, message_title, not_level)

    def progressbar(self, progress, message_text="", not_level_reverse="info"):
        """Creates and updates progress bar. Only used with Pyplan UI
           Ex.
               pp.progressbar(20, "Step 1","info")
               pp.progressbar(100, "Complete!","success")
        """

        if self.model and self.model.ws:
            not_level = ws_settings.NOTIFICATION_LEVEL_CHOICES_REVERSE[
                not_level_reverse] if not_level_reverse in ws_settings.NOTIFICATION_LEVEL_CHOICES_REVERSE else ws_settings.NOTIFICATION_LEVEL_INFO
            self.model.ws.progressbar(progress, message_text, not_level)

    def create_report(self, reportItems, reportIndexName="Report index", reportIndex=None):
        """Concatenates the reportItems dic dataArrays along the reportIndex dimension

           reportItems: dict or list with datarrays to concat (must have the same structure)
           reportIndexName: Name of the new ReportIndex dimension
           reportIndex: Overwrite ReportIndex dimension
           Ex.
               pp.create_report(reportItems={"Demand":demand, "Product Stock":stock}, reportIndexName="New Report")
        """

        if isinstance(reportItems, dict):
            report_index = list(reportItems)
            report_values = list(reportItems.values())
            _titles = [str(xx.name) for xx in report_values]
            _index = pd.Index(report_index, name=reportIndexName)
            return xr.concat(report_values, _index)
        else:
            _titles = [str(xx.name) for xx in reportItems]
            _index = None
            if reportIndex is None:
                _index = pd.Index(_titles, name=reportIndexName)
            else:
                _index = reportIndex
            return xr.concat(reportItems, _index)

    def pandas_from_dataarray(self, dataarray):
        """Create dataframe pandas from datarray with n dimensions

           Ex.
               pp.pandas_from_dataarray(dataArrayNode)
        """

        return dataarray.stack(z=dataarray.dims).to_dataframe("value")

    def pandas_from_access(self):
        """Class to manage access databases
        """

        return Pandas_from_acc()

    def __generate_pkl_from_excel(self, workbook, filepath, targetDir, maxFileSizeMB=None, flagFilename='flag.tmp'):
        """Generates compressed pickle from excel file

           workbook: openpyxl workbook
           filepath: full file path
           targetDir: path where pickles will be stored
           maxFileSizeMB: file size limit in megabytes
           flagFilename: name of temporary flag file
        """

        optimizable_templates = ['.xlsx', '.xlsm', '.xlsb']
        _, ext = os.path.splitext(filepath)

        # Generate pickle for selected file types if its size is below max limit
        if ext in optimizable_templates and (maxFileSizeMB is None or os.stat(filepath).st_size/1024/1024 <= maxFileSizeMB):
            if not os.path.isdir(targetDir):
                os.mkdir(targetDir)

            # When first user runs optimization, creates flag file that gets deleted after whole optimization is done
            # If another user wants to read the Excel file while the optimization is running, the flag file will be present
            flag_filepath = os.path.join(targetDir, flagFilename)
            with open(flag_filepath, 'w'):
                pass

            try:
                for item in workbook.defined_names.definedName:
                    try:
                        if not item.is_external and item.type == 'RANGE' and item.attr_text and '!$' in item.attr_text:
                            target_filepath = os.path.join(
                                targetDir, f'{item.name}.pkl')
                            if os.path.isfile(target_filepath):
                                os.remove(target_filepath)

                            dests = item.destinations
                            for title, coord in dests:
                                if title in workbook:
                                    ws = workbook[title]
                                    rangeToRead = ws[coord]
                                    if not isinstance(rangeToRead, tuple):
                                        rangeToRead = ((rangeToRead,),)

                                    cols = []
                                    values = []
                                    for index, row in enumerate(rangeToRead):
                                        if index == 0:
                                            cols = [str(c.value) for c in row]
                                        else:
                                            values.append(
                                                [c.value for c in row])
                                    nn = 0
                                    _finalCols = []
                                    for _col in cols:
                                        if _col is None:
                                            _finalCols.append(
                                                f'Unnamed{str(nn)}')
                                            nn += 1
                                        else:
                                            _finalCols.append(_col)
                                    df = pd.DataFrame(
                                        values, columns=_finalCols).dropna(how='all')
                                    df.to_pickle(target_filepath,
                                                 compression='gzip')
                    except Exception as e:
                        print(
                            f"Could not generate pkl for range '{item.name}'. Error: {e}")
            finally:
                os.remove(flag_filepath)

    def __remove_old_file(self, filepath, maxElapsedMinutes=1):
        """Deletes file if its modified date is older than current date minus maxElapsedMinutes
        """

        if os.path.isfile(filepath):
            # Dates are expressed in seconds since epoch (floats)
            modified_date = os.path.getmtime(filepath)
            min_modified_date = time.time() - (maxElapsedMinutes * 60)
            if modified_date < min_modified_date:
                os.remove(filepath)

    def __read_pickle_df(self, filepath, indexes=None):
        """Loads dataframe from pickled file
        """

        df = pd.read_pickle(filepath, compression='gzip')
        if not indexes is None:
            df.set_index(indexes, inplace=True)
        return df

    def get_nested_lists_shape(self, lst, shape=()):
        """Returns a tuple with the shape of nested lists similarly to numpy's shape.

           lst: the nested list
           shape: the shape up to the current recursion depth
        """

        if not isinstance(lst, list):
            # base case
            return shape

        # peek ahead and assure all lists in the next depth have the same length
        if isinstance(lst[0], list):
            l = len(lst[0])
            if not all(len(item) == l for item in lst):
                raise ValueError('Not all lists have the same length')

        shape += (len(lst), )

        # recurse
        shape = self.get_nested_lists_shape(lst[0], shape)

        return shape

    def __concat_dataarrays_over_one_dim(self, valuesList, dim):
        """Concatenates Xarray DataArrays along a new dimension, broadcasting by all possible
        dimensions

           valuesList: list of DataArrays, int, str, float. At least one of them must be DataArray
           dim: Pandas Index with same length as valuesList
        """

        # Error handling
        if not isinstance(valuesList, list):
            raise ValueError('valuesList must be a list')
        if not any([isinstance(v, xr.DataArray) for v in valuesList]):
            raise ValueError(
                'At least one of the objects in valuesList must be a Xarray DataArray')
        if not isinstance(dim, pd.Index):
            raise ValueError('dim must be a pandas Index')
        valuesListShape = self.get_nested_lists_shape(valuesList)
        dimShape = (len(dim.values),)
        if valuesListShape != dimShape:
            raise ValueError(
                f'Shape of valuesList {valuesListShape} is not equal to shape of dim {dimShape}')

        # Get all possible dimensions
        all_dims_names, all_dims_indexes = [], {}
        for v in valuesList:
            if isinstance(v, xr.DataArray):
                dims_v = v.dims
                indexes_v = v.indexes
                for d in dims_v:
                    if d not in all_dims_names:
                        all_dims_names.append(d)
                        all_dims_indexes.update({d: indexes_v[d]})

        newValuesList = []
        for v in valuesList:
            # Add dimensions not present in original DataArray
            if isinstance(v, xr.DataArray):
                dims_v = v.dims
                if not all(d in dims_v for d in all_dims_names):
                    for d in all_dims_names:
                        if d not in dims_v:
                            v = v.expand_dims({d: all_dims_indexes[d].values})
            # When value is a scalar (str, int, float, usually)
            else:
                v = xr.DataArray(v, list(all_dims_indexes.values()))

            newValuesList.append(v)

        return xr.concat(newValuesList, dim=dim)

    def concat_dataarrays(self, valuesList, dim):
        """Concatenates Xarray DataArrays along one or two new dimensions, broadcasting
        by all possible dimensions

           valuesList: list or list of lists of DataArrays, int, str, float. At least one
           of them must be a DataArray object
           dim: Pandas Index or list of Pandas Indexes with same shape as valuesList

           Ex.
               pp.concat_dataarrays(
                   valuesList=[node1, node2, node3],
                   dim=three_items_index)
               pp.concat_dataarrays(
                   valuesList=['String Example', node2, 0],
                   dim=three_items_index)
               pp.concat_dataarrays(
                   valuesList=[[node1, node2, node3], [node4, node5, node6]],
                   dim=[two_items_index, three_items_index])
        """

        valuesListShape = self.get_nested_lists_shape(valuesList)
        if isinstance(dim, list):
            # Error handling
            # Check length
            if len(dim) > 2:
                raise ValueError("Can only concat along 1 or 2 dimensions")
            if len(valuesListShape) == 1:
                raise ValueError(
                    "valuesList must be list of lists if dim is a list")

            # Ensure shapes are the same
            if valuesListShape[0] == 1:
                # to make it comparable to dimShape
                valuesListShape = (valuesListShape[1],)
            dimShape = tuple(len(d) for d in dim)
            if not (valuesListShape == dimShape):
                raise ValueError(
                    f"Shape of valuesList {valuesListShape} is not equal to shape of dim {dimShape}")

            if len(dimShape) == 2:
                # Broadcast and concat along each dim
                das = []
                for lst in valuesList:
                    da = self.__concat_dataarrays_over_one_dim(
                        valuesList=lst, dim=dim[1])
                    das.append(da)
                return self.__concat_dataarrays_over_one_dim(valuesList=das, dim=dim[0])
            else:
                return self.__concat_dataarrays_over_one_dim(valuesList=valuesList[0], dim=dim[0])
        else:
            return self.__concat_dataarrays_over_one_dim(valuesList=valuesList, dim=dim)


class Selector(object):
    """ Class to manage UI Pyplan selectors.
    """

    SERIALIZABLE_PROPERTIES = ['options', 'selected', 'multiselect']

    def __init__(self, options, selected, multiselect=False):
        """ Create UI Pyplan selector for desicion nodes
        Params:
            options: List or pd.Index with available values that can be selected 
            selected: current selected index value's
            multiselect: True to allow multiple selection
        """
        self._options = options
        self._multiselect = multiselect
        self.selected = selected

    @property
    def value(self):
        if self.multiselect:
            return [self.options[i] for i in self.selected]
        else:
            return self.options[self.selected]

    @property
    def options(self):
        return self._options

    @property
    def multiselect(self):
        return self._multiselect

    @property
    def selected(self):
        res = None
        if self.multiselect:
            res = []
            for nn in self._selected:
                if nn < len(self._options):
                    res.append(nn)
            if len(res) == 0:
                res = list(range(len(self._options)))
        else:
            res = self._selected if self._selected < len(self._options) else 0

        return res

    @selected.setter
    def selected(self, value):
        if self.multiselect:
            if value is None:
                self._selected = []
            elif isinstance(value, list):
                self._selected = value
            else:
                self._selected = [value]
        else:
            if isinstance(value, list):
                self._selected = value[0]
            else:
                self._selected = value

    def toObj(self):
        res = dict()
        for k in Selector.SERIALIZABLE_PROPERTIES:
            if hasattr(self, k):
                if k == "options" and isinstance(getattr(self, k), pd.Index):
                    res[k] = getattr(self, k).tolist()
                else:
                    res[k] = getattr(self, k)
        return res

    def isSameValue(self, value):
        if self.multiselect and isinstance(self.selected, list) and isinstance(value, list):
            l1 = self.selected.copy()
            l2 = value.copy()
            l1.sort()
            l2.sort()
            return l1 == l2
        else:
            return self.selected == value

    def generateDefinition(self, definition, value):

        if self.multiselect:
            if not isinstance(value, list):
                if value is None:
                    value = 0
                value = list(value)
            elif len(value) == 0:
                value = list(range(len(self.options)))
        newPos = str(value)

        reg = r'(?:[^\]\[,]+|\[[^\]\[]+\])'
        groups = re.findall(reg, definition)
        if len(groups) > 2:
            if not str(groups[-1]) in ["False)", "True)", "multiselect=False)", "multiselect=True)"]:
                groups.append("False)")
            newDef = ""
            for nn in range(len(groups)-2):
                newDef += groups[nn]
            newDef = f"{newDef},{newPos},{groups[-1]}"
            return newDef
        return None


class Pandas_from_acc():
    """Class that allows to read access files with pandas

    EXAMPLES OF USE:

    # Listing the tables.
    for tbl in pandas_from_access.list_tables("my.mdb"):
        print(tbl)

    # Read a small table.
    df = pandas_from_access.read_table("my.mdb", "MyTable")

    # Read a huge table.
    accumulator = []
    for chunk in pandas_from_access.read_table("my.mdb", "MyTable", chunksize=10000):
        accumulator.append(f(chunk))
    """

    TABLE_RE = re.compile("CREATE TABLE \[([a-zA-Z_0-9 ]+)\]\s+\((.*?\));",
                          re.MULTILINE | re.DOTALL)
    DEF_RE = re.compile("\s*\[(\w+)\]\s*(.*?),")

    @classmethod
    def list_tables(cls, rdb_file, encoding="latin-1"):
        """
        :param rdb_file: The MS Access database file.
        :param encoding: The content encoding of the output. I assume `latin-1`
            because so many of MS files have that encoding. But, MDBTools may
            actually be UTF-8.
        :return: A list of the tables in a given database.
        """
        tables = cls.__get_tables(rdb_file, encoding)
        return [table for table, _ in tables]

    @classmethod
    def read_schema(cls, rdb_file, encoding='utf8'):
        """
        :param rdb_file: The MS Access database file.
        :param encoding: The schema encoding. I'm almost positive that MDBTools
            spits out UTF-8, exclusively.
        :return: a dictionary of table -> column -> access_data_type
        """
        output = subprocess.check_output(['mdb-schema', rdb_file])
        lines = output.decode(encoding).splitlines()
        schema_ddl = "\n".join(l for l in lines if l and not l.startswith('-'))
        tables = cls.__get_tables(rdb_file, encoding)
        schema = {}
        for table, defs in tables:
            schema[table] = cls.__extract_defs(defs)

        return schema

    @classmethod
    def to_pandas_schema(cls, schema, implicit_string=True):
        """
        :param schema: the output of `read_schema`
        :param implicit_string: mark strings and unknown dtypes as `np.str_`.
        :return: a dictionary of table -> column -> np.dtype
        """
        pd_schema = {}
        for tbl, defs in schema.items():
            pd_schema[tbl] = None
            sub_schema = {}
            for column, data_type in defs.items():
                dtype = cls.__extract_dtype(data_type)
                if dtype is not None:
                    sub_schema[column] = dtype
                elif implicit_string:
                    sub_schema[column] = np.str_
            pd_schema[tbl] = sub_schema
        return pd_schema

    @classmethod
    def read_table(cls, rdb_file, table_name, *args, **kwargs):
        """
        Read a MS Access database as a Pandas DataFrame.

        Unless you set `converters_from_schema=False`, this function assumes you
        want to infer the schema from the Access database's schema. This sets the
        `dtype` argument of `read_csv`, which makes things much faster, in most
        cases. If you set the `dtype` keyword argument also, it overrides
        inferences. The `schema_encoding keyword argument passes through to
        `read_schema`. The `implicit_string` argument passes through to
        `to_pandas_schema`.

        I recommend setting `chunksize=k`, where k is some reasonable number of
        rows. This is a simple interface, that doesn't do basic things like
        counting the number of rows ahead of time. You may inadvertently start
        reading a 100TB file into memory. (Although, being a MS product, I assume
        the Access format breaks after 2^32 bytes -- har, har.)

        :param rdb_file: The MS Access database file.
        :param table_name: The name of the table to process.
        :param args: positional arguments passed to `pd.read_csv`
        :param kwargs: keyword arguments passed to `pd.read_csv`
        :return: a pandas `DataFrame` (or, `TextFileReader` if you set
            `chunksize=k`)
        """
        if kwargs.pop('converters_from_schema', True):
            specified_dtypes = kwargs.pop('dtype', {})
            schema_encoding = kwargs.pop('schema_encoding', 'utf8')
            schemas = cls.to_pandas_schema(cls.read_schema(rdb_file, schema_encoding),
                                           kwargs.pop('implicit_string', True))
            dtypes = schemas[table_name]
            dtypes.update(specified_dtypes)
            if dtypes != {}:
                kwargs['dtype'] = dtypes

        cmd = ['mdb-export', rdb_file, table_name]
        proc = subprocess.Popen(cmd, stdout=subprocess.PIPE)
        return pd.read_csv(proc.stdout, *args, **kwargs)

    # private class methods
    @classmethod
    def __get_tables(cls, rdb_file, encoding='utf8'):
        output = subprocess.check_output(['mdb-schema', rdb_file])
        lines = output.decode(encoding).splitlines()
        schema_ddl = "\n".join(l for l in lines if l and not l.startswith('-'))
        return Pandas_from_acc.TABLE_RE.findall(schema_ddl)

    @classmethod
    def __extract_dtype(cls, data_type):
        # Note, this list is surely incomplete. But, I only had one .mdb file
        # at the time of creation. If you see a new data-type, patch-pull or just
        # open an issue.
        data_type = data_type.lower()
        if data_type.startswith('double'):
            return np.float_
        elif data_type.startswith('long'):
            return np.float_
        elif data_type.startswith('bool'):
            return np.bool_
        elif data_type.startswith('text') or data_type.startswith('memo'):
            return np.str_
        elif data_type.startswith('ole'):
            return np.bytes_
        else:
            return None

    @classmethod
    def __extract_defs(cls, defs_str):
        defs = {}
        lines = defs_str.splitlines()
        for line in lines:
            m = Pandas_from_acc.DEF_RE.match(line)
            if m:
                defs[m.group(1)] = m.group(2)
        return defs
